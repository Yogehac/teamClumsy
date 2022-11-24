# Importing libraries

import os.path
import openpyxl
import pandas as pd


# Initializing folder path
# Path for raw excel
# df = "D:/IA/new/excel/quot.xlsx"




# This class takes in unprocessed excel file and converts into processed CSV 
class single_xlsx_processing:
    
    # Takes in file with its path as input
    def __init__(self, df):
        
        self.wb = openpyxl.load_workbook(df)
        self.sheet = self.wb.active
        
        # Output file path and extension for comma removed excel
        self.op_comma_removed_file = os.path.splitext(df)[0] + "_comma_removed" + os.path.splitext(df)[1]
        
        # Output file path and extension for processed excel
        self.opfile = os.path.splitext(df)[0] + "_processed" + os.path.splitext(df)[1]
        
        # Output file path and extension for processed CSV
        self.opcsvfile = os.path.splitext(df)[0] + ".csv"
        
        self.remove_comma()
        self.unmerge_cells()
        self.del_empty_cell()
        self.xlsx_to_CSV()
        
    def remove_comma(self):
        for row in self.sheet.iter_rows():
            for cell in row:
                if(cell.value != None):
                    if (type(cell.value) == str):
                        #print(cell.value)
                        r = cell.value.replace(",", "-")
                        cell.value = r
                
        self.wb.save(self.op_comma_removed_file)
        
    # Unmerges merged cells in excel 
    def unmerge_cells(self):
        self.wb1 = openpyxl.load_workbook(self.op_comma_removed_file)
        self.sheet1 = self.wb1.active
        for items in sorted(self.sheet1.merged_cell_ranges):
            
            self.sheet1.unmerge_cells(str(items))
      
        self.wb1.save(self.opfile)
        
    # Removes all the empty cells and columns in the dataset
    def del_empty_cell(self):
        
        df = pd.read_excel(self.opfile)
        df = df.apply(lambda x: pd.Series(x.dropna()),
                      axis = 1).fillna(' ')
        df = df.to_excel(self.opfile)
    
    # Converts excel to CSV format    
    def xlsx_to_CSV(self):
        
        df = pd.read_excel(self.opfile, index_col = 0)
        
        df.to_csv(self.opcsvfile[:-4]+'.csv',
                  encoding ='utf-8',
                  index = False)
        




# single_xlsx_processing(df)
